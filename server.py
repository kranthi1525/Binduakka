"""
Asset Management Server
Run:  python server.py
Team: http://<YOUR-IP>:5000   (shown when server starts)

All requests are saved to requests.json on disk so they
survive server restarts and are visible from any machine.
"""

from flask import Flask, jsonify, request, session, send_from_directory
import openpyxl, os, json
from datetime import datetime

app = Flask(__name__)
app.secret_key = "zf_asset_mgmt_2024_secret"

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE    = os.path.join(BASE_DIR, "Asset_Tracking.xlsx")
REQUESTS_FILE = os.path.join(BASE_DIR, "requests.json")

# ── Admin credentials ──────────────────────────────────────────────────────────
ADMIN_USERS = {
    "bindumadhavi": "Welcome@123",
    "nagarjuna":    "Welcome@123",
}

# ── In-memory asset store ─────────────────────────────────────────────────────
sw_assets  = []
hw_assets  = []
id_counter = {"sw": 1, "hw": 1, "req": 1}

# ── Requests persisted to disk ────────────────────────────────────────────────
def load_requests():
    """Load requests.json from disk into memory."""
    global id_counter
    if not os.path.exists(REQUESTS_FILE):
        return [], []
    try:
        with open(REQUESTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        pending = data.get("pending", [])
        history = data.get("history", [])
        # Advance req counter past existing IDs
        all_ids = [r["id"] for r in pending + history]
        for rid in all_ids:
            try:
                num = int(rid.replace("REQ-", ""))
                if num >= id_counter["req"]:
                    id_counter["req"] = num + 1
            except:
                pass
        print(f"[INFO] Loaded {len(pending)} pending, {len(history)} history from requests.json")
        return pending, history
    except Exception as e:
        print(f"[WARN] Could not read requests.json: {e}")
        return [], []

def save_requests(pending, history):
    """Save current pending + history to requests.json."""
    try:
        with open(REQUESTS_FILE, "w", encoding="utf-8") as f:
            json.dump({"pending": pending, "history": history}, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"[WARN] Could not save requests.json: {e}")

# Load on startup
requests_q, history_log = load_requests()

# ── Parse Excel ───────────────────────────────────────────────────────────────
def load_excel():
    global sw_assets, hw_assets
    if not os.path.exists(EXCEL_FILE):
        print(f"[WARN] Excel not found: {EXCEL_FILE}")
        return
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    sw_assets = []
    if "SW_Licenses" in wb.sheetnames:
        ws  = wb["SW_Licenses"]
        hdrs = [c.value for c in ws[1]]
        ctr  = 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            r    = dict(zip(hdrs, row))
            name = str(r.get("License Name") or "").strip()
            if not name: continue
            sw_assets.append({
                "id":         f"sw_{ctr:04d}",
                "sheet":      "sw",
                "name":       name,
                "version":    str(r.get("Version") or ""),
                "license_id": str(r.get("LicenseID") or ""),
                "device":     str(r.get("ConfiguredDevice") or ""),
                "serial":     str(r.get("Device Serial Number") or ""),
                "po":         str(r.get("PO/AWB") or ""),
                "model":      str(r.get("License Model") or ""),
                "used_for":   str(r.get("Used on/for") or ""),
                "project":    str(r.get("Which Project") or ""),
                "assigned":   str(r.get("Currently assigned to") or "").strip(),
                "expiry":     str(r.get("Expriry Data") or ""),
                "used":       str(r.get("Used/Unused") or ""),
                "region":     str(r.get("Region") or ""),
                "remarks":    str(r.get("Remarks") or ""),
            })
            ctr += 1
        id_counter["sw"] = ctr

    hw_assets = []
    if "HW" in wb.sheetnames:
        ws   = wb["HW"]
        hdrs = [c.value for c in ws[1]]
        ctr  = 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            r       = dict(zip(hdrs, row))
            name    = str(r.get("Device Name") or "").strip()
            hw_type = str(r.get("HW Type") or "").strip()
            if not name and not hw_type: continue
            hw_assets.append({
                "id":          f"hw_{ctr:04d}",
                "sheet":       "hw",
                "hw_type":     hw_type,
                "name":        name,
                "description": str(r.get("Description") or ""),
                "serial":      str(r.get("Serial No") or ""),
                "po":          str(r.get("PO") or ""),
                "awb":         str(r.get("AWB") or ""),
                "project":     str(r.get("Which Project") or ""),
                "assigned":    str(r.get("Assigned to") or "").strip(),
                "comments":    str(r.get("Comments") or ""),
                "used":        str(r.get("Used/ Unused") or ""),
                "region":      str(r.get("Region") or ""),
                "remarks":     str(r.get("Remarks") or ""),
            })
            ctr += 1
        id_counter["hw"] = ctr

    # Apply any approved reassignments from history
    for req in history_log:
        if req.get("status") == "approved":
            for lst in (sw_assets, hw_assets):
                for a in lst:
                    if a["id"] == req.get("asset_id"):
                        a["assigned"] = req["new_assignee"]

    print(f"[INFO] Loaded {len(sw_assets)} SW licenses, {len(hw_assets)} HW assets.")

# ── Serve dashboard ────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "dashboard.html")

# ── Auth ───────────────────────────────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
def login():
    d = request.get_json()
    u = (d.get("username") or "").strip().lower()
    p = (d.get("password") or "")
    for key, pwd in ADMIN_USERS.items():
        if u == key.lower() and p == pwd:
            session["admin"] = True
            return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Invalid credentials"}), 401

@app.route("/api/logout", methods=["POST"])
def logout():
    session.pop("admin", None)
    return jsonify({"ok": True})

@app.route("/api/me")
def me():
    return jsonify({"admin": bool(session.get("admin"))})

# ── Assets ─────────────────────────────────────────────────────────────────────
@app.route("/api/assets")
def get_assets():
    t = request.args.get("type", "sw")
    return jsonify(sw_assets if t == "sw" else hw_assets)

@app.route("/api/stats")
def stats():
    return jsonify({
        "sw":      len(sw_assets),
        "hw":      len(hw_assets),
        "pending": len(requests_q),
        "history": len(history_log),
    })

# ── Admin: Add asset ───────────────────────────────────────────────────────────
@app.route("/api/asset", methods=["POST"])
def add_asset():
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 403
    d     = request.get_json()
    sheet = d.get("sheet", "sw")
    if sheet == "sw":
        asset = {
            "id":         f"sw_{id_counter['sw']:04d}",
            "sheet":      "sw",
            "name":       (d.get("name") or "").strip(),
            "version":    d.get("version",""),
            "license_id": d.get("license_id",""),
            "device":     d.get("device",""),
            "serial":     d.get("serial",""),
            "po":         d.get("po",""),
            "model":      d.get("model",""),
            "used_for":   d.get("used_for",""),
            "project":    d.get("project",""),
            "assigned":   (d.get("assigned") or "").strip(),
            "expiry":     d.get("expiry",""),
            "used":       d.get("used",""),
            "region":     d.get("region",""),
            "remarks":    d.get("remarks",""),
        }
        if not asset["name"]:
            return jsonify({"error": "License name is required"}), 400
        sw_assets.append(asset)
        id_counter["sw"] += 1
    else:
        asset = {
            "id":          f"hw_{id_counter['hw']:04d}",
            "sheet":       "hw",
            "hw_type":     (d.get("hw_type") or "").strip(),
            "name":        (d.get("name") or "").strip(),
            "description": d.get("description",""),
            "serial":      d.get("serial",""),
            "po":          d.get("po",""),
            "awb":         d.get("awb",""),
            "project":     d.get("project",""),
            "assigned":    (d.get("assigned") or "").strip(),
            "comments":    d.get("comments",""),
            "used":        d.get("used",""),
            "region":      d.get("region",""),
            "remarks":     d.get("remarks",""),
        }
        if not asset["name"] and not asset["hw_type"]:
            return jsonify({"error": "Device name or HW type is required"}), 400
        hw_assets.append(asset)
        id_counter["hw"] += 1
    return jsonify({"ok": True, "asset": asset})

# ── Admin: Delete asset ────────────────────────────────────────────────────────
@app.route("/api/asset/<asset_id>", methods=["DELETE"])
def delete_asset(asset_id):
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 403
    for lst in (sw_assets, hw_assets):
        for a in lst:
            if a["id"] == asset_id:
                lst.remove(a)
                return jsonify({"ok": True})
    return jsonify({"error": "Asset not found"}), 404

# ── Submit reassign request (ANY team member, no auth needed) ──────────────────
@app.route("/api/request", methods=["POST"])
def submit_request():
    d = request.get_json()
    for field in ("asset_id","asset_name","asset_sheet","current_assignee","new_assignee","requested_by"):
        if not (d.get(field) or "").strip():
            return jsonify({"error": f"Missing field: {field}"}), 400
    req = {
        "id":               f"REQ-{id_counter['req']:04d}",
        "asset_id":         d["asset_id"],
        "asset_name":       d["asset_name"],
        "asset_sheet":      d["asset_sheet"],
        "current_assignee": d["current_assignee"],
        "new_assignee":     d["new_assignee"],
        "requested_by":     d["requested_by"],
        "requester_email":  d.get("requester_email",""),
        "reason":           d.get("reason",""),
        "created_at":       datetime.now().strftime("%d-%m-%Y %H:%M"),
        "status":           "pending",
        "admin_note":       "",
        "decided_at":       "",
    }
    id_counter["req"] += 1
    requests_q.append(req)
    save_requests(requests_q, history_log)   # ← persist to disk immediately
    return jsonify({"ok": True, "request_id": req["id"]})

# ── Pending + History (admin only) ────────────────────────────────────────────
@app.route("/api/requests/pending")
def pending():
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 403
    return jsonify(requests_q)

@app.route("/api/requests/history")
def history():
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 403
    return jsonify(history_log)

# ── Admin: Approve / Reject ───────────────────────────────────────────────────
@app.route("/api/request/<req_id>/decide", methods=["POST"])
def decide(req_id):
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 403
    d        = request.get_json()
    decision = d.get("decision")
    note     = d.get("admin_note","")
    req      = next((r for r in requests_q if r["id"] == req_id), None)
    if not req:
        return jsonify({"error": "Request not found"}), 404
    req["status"]     = decision
    req["admin_note"] = note
    req["decided_at"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    if decision == "approved":
        for lst in (sw_assets, hw_assets):
            for a in lst:
                if a["id"] == req["asset_id"]:
                    a["assigned"] = req["new_assignee"]
                    break
    requests_q.remove(req)
    history_log.insert(0, req)
    save_requests(requests_q, history_log)   # ← persist to disk immediately
    return jsonify({"ok": True, "req": req})

# ── Admin: reload Excel ────────────────────────────────────────────────────────
@app.route("/api/reload", methods=["POST"])
def reload_excel():
    if not session.get("admin"):
        return jsonify({"error": "Unauthorized"}), 403
    load_excel()
    return jsonify({"ok": True, "sw": len(sw_assets), "hw": len(hw_assets)})

# ── Startup ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    load_excel()
    import socket
    try:
        ip = socket.gethostbyname(socket.gethostname())
    except:
        ip = "localhost"
    print(f"\n{'='*60}")
    print(f"  Asset Management Dashboard is RUNNING")
    print(f"  Your PC  : http://localhost:5000")
    print(f"  Team URL : http://{ip}:5000")
    print(f"  Share the Team URL with everyone on your network.")
    print(f"{'='*60}\n")
    app.run(host="0.0.0.0", port=5000, debug=False)
